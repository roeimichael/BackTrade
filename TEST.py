import pandas as pd
import os
from openpyxl.styles import PatternFill
from openpyxl import load_workbook


def color_columns(file_path):
    # Load the Excel file
    wb = load_workbook(file_path)
    sheet = wb.active

    # Create a dictionary of different colors for the column headers based on file name
    colors = {'5': '0000FF', '20': '00FF00', '50': 'FFFF00', '100': 'FF00FF'}

    # Iterate through all columns
    for col in sheet.iter_cols():
        # Get the file ending from the column header
        file_ending = col[0].value.split('_')[-1]
        color = colors.get(file_ending)

        # Apply the color for the current file to the column headers
        for cell in col:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    wb.save(file_path)


import pandas as pd
import os


def combine_csv_files(folder_path):
    # Initialize empty list to store all data
    dataframes = []
    # Iterate through all CSV files in the folder
    for csv_file in os.listdir(folder_path):
        if csv_file.endswith('.csv'):
            file_path = os.path.join(folder_path, csv_file)
            # Read the CSV file and only select the first 3 columns
            temp_df = pd.read_csv(file_path, usecols=[0, 1, 2])
            # Append the data to the list
            dataframes.append((temp_df, csv_file))
    # Combine the dataframes using pd.concat()
    df = pd.concat([x[0] for x in dataframes], axis=1, keys=[x[1] for x in dataframes])
    # Save the combined dataframe to a new CSV file
    df.to_csv(os.path.join(folder_path, 'final.csv'), index=False)


combine_csv_files('./data/PrecisionTesting/2.5/')
